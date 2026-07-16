#!/usr/bin/env python3
"""
map_headers.py — the ONE place the LLM is involved. AI maps, code moves numbers.

Given a supplier quote, this proposes the mapping.json that process_quote.py then
uses to move values deterministically. The model only ever sees the header
candidate rows plus a few sample data rows — never the full rate table — and it
only ever returns column-name -> field associations. It cannot emit a rate, EAC
or meter point into the pipeline; process_quote.py copies those from the source.

Backend endpoints this backs:
  * /inspect  -> inspect_file()          (pure, no network)
  * /map      -> propose_mapping()        (the single LLM call)

LLM transport is env-swappable so the OUTSTANDING Vercel AI Gateway question
changes nothing here:
  ANTHROPIC_API_KEY   required for a live call
  ANTHROPIC_BASE_URL  optional — point at the Vercel AI Gateway to route
                      through it (spend monitoring); unset = direct Anthropic API
  ANTHROPIC_MODEL     optional — default 'claude-sonnet-5'

CLI:
    python3 map_headers.py SOURCE [--supplier NAME] [--sample-rows N]
                                  [--dry-run] [--out mapping.json]
    --dry-run prints the inspection + the EXACT request that would be sent
    (proving the payload is headers + samples only) and makes no network call.
"""
import argparse
import hashlib
import json
import os
import re
import sys

from rye_quote_core import TARGET_FIELDS

# Keywords that hint a row is the real header row (not a metadata/branding row).
HEADER_HINTS = [
    "mpan", "mprn", "mpxn", "meter", "supply number", "eac", "aq", "consumption",
    "rate", "standing", "day", "night", "peak", "unit", "kva", "capacity",
    "duos", "network", "site", "address", "start", "charge",
]

DEFAULT_MODEL = "claude-sonnet-5"


# --- /inspect : pure, no network -------------------------------------------

def _cell(v):
    return "" if v is None else str(v).strip()


_NUMISH = re.compile(r"^[-+]?[\d,]*\.?\d+$")


def _value_like(cell):
    """True if a cell reads as a value (number/date/id) rather than a label.

    A real header cell is a short text label. Data and summary-block cells look
    like values: pure numbers, or anything with a run of 2+ digits (dates, IDs,
    long figures). Used to reject 'label: value' summary rows as headers.
    """
    c = cell.strip()
    if not c:
        return False
    if _NUMISH.match(c.replace("£", "").replace("%", "").replace(" ", "")):
        return True
    return bool(re.search(r"\d{2,}", c))


def _score_header_row(grid, idx):
    """Score row `idx` on how much it looks like a real table header.

    A header (a) is mostly short text labels, not values; (b) often carries
    energy keywords; and crucially (c) is followed by a wide, consistent block
    of data that populates its columns. Point (c) is what stops a keyword-rich
    'Quote summary' block (label/value pairs, little below it) from beating the
    actual 'Rates breakdown' header several rows further down.
    """
    cells = grid[idx]
    cols = [j for j, c in enumerate(cells) if c.strip()]
    n = len(cols)
    if n < 2:
        return 0
    labels = [cells[j].strip() for j in cols]
    if sum(1 for c in labels if _value_like(c)) / n >= 0.4:
        return 0  # too many value-like cells => a data/summary row, not a header
    kw = sum(1 for c in labels if any(h in c.lower() for h in HEADER_HINTS))
    # Data-below overlap: how well the next few non-empty rows fill these columns.
    overlaps = []
    for k in range(idx + 1, min(idx + 7, len(grid))):
        r = grid[k]
        if not any(x.strip() for x in r):
            continue
        overlaps.append(sum(1 for j in cols if j < len(r) and r[j].strip()) / n)
        if len(overlaps) >= 4:
            break
    overlap = sum(overlaps) / len(overlaps) if overlaps else 0.0
    return n * (0.5 + overlap) + 3 * kw


def _sheet_grid(path, sheet, max_rows):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        grid = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            grid.append([_cell(c) for c in row])
        return grid
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [[_cell(c) for c in r] for r in list(__import__("csv").reader(f))[:max_rows]]


def inspect_file(path, max_rows=40):
    """Sheet names, first ~max_rows rows, and ranked header-row candidates.

    This is exactly what the /inspect endpoint returns and is the ONLY view of
    the file the mapping step gets beyond a handful of sample rows. Scans enough
    rows to reach a header that sits below a summary block (e.g. Octopus
    multisite quotes put the rates table ~20 rows down).
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet_names = wb.sheetnames
    else:
        sheet_names = [None]  # single logical sheet for csv

    sheets = []
    for name in sheet_names:
        grid = _sheet_grid(path, name, max_rows)
        scored = sorted(
            ((_score_header_row(grid, idx), idx + 1) for idx in range(len(grid))),
            key=lambda t: (t[0], -t[1]),  # ties: prefer the earlier (topmost) row
            reverse=True,
        )
        candidates = [row for score, row in scored[:3] if score > 0]
        best = candidates[0] if candidates else 1
        sheets.append({
            "name": name,
            "header_row_candidates": candidates,
            "header_row_best_guess": best,
            "headers": grid[best - 1] if best - 1 < len(grid) else [],
            "first_rows": grid,
        })
    return {"path": os.path.basename(path), "sheets": sheets}


# --- layout fingerprint : the learned-mappings cache key -------------------

def _norm_header(cell):
    """Normalise a header label so cosmetic differences don't fork the cache.

    Lower-cases, strips currency/percent decoration, and collapses any run of
    non-alphanumerics to a single space. 'Unit Rate (p/kWh)' and 'unit rate'
    both reduce to 'unit rate' — the same layout, the same fingerprint.
    """
    s = str(cell).strip().lower()
    s = re.sub(r"[£$%]", "", s)
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return s


def header_signature(inspection):
    """A stable, human-readable signature of a file's layout.

    Built ONLY from each sheet's best-guess header labels (never data values),
    so the same supplier layout filled with a different client's numbers yields
    the same signature. Sheet order is preserved; sheet *names* are excluded
    because they often carry the client/date and would over-fork the cache.
    """
    parts = []
    for s in inspection.get("sheets", []):
        labels = [_norm_header(h) for h in s.get("headers", []) if str(h).strip()]
        labels = [l for l in labels if l]
        parts.append("|".join(labels))
    return "||".join(parts)


def layout_fingerprint(inspection):
    """Hash of the header signature — the `layout_fingerprint` cache column.

    Deterministic and collision-resistant; paired with `supplier` it is the
    unique key of the `supplier_mappings` cache. Same layout in, same hash out,
    so a repeat supplier layout is served from cache and skips the LLM.
    """
    sig = header_signature(inspection)
    return hashlib.sha256(sig.encode("utf-8")).hexdigest()[:16]


# --- sample values : the confirm/override screen ---------------------------

def _mapped_headers(mapping):
    """Flatten a mapping's `columns` into {target_field: source_header}.

    Column specs are null | "H" | {"single": "H"} | {"split": "H"}; the sentinel
    "__none__" (used by the prompt to say 'no such band') is treated as unmapped.
    """
    out = {}
    for field, spec in (mapping.get("columns") or {}).items():
        header = None
        if isinstance(spec, str):
            header = spec
        elif isinstance(spec, dict):
            header = spec.get("single") or spec.get("split")
        if header and header != "__none__":
            out[field] = header
    return out


def sample_values(inspection, mapping, max_samples=3):
    """For each mapped field, the source header + a few example cell values.

    Powers the human confirm/override screen: the reviewer sees which column
    feeds each field and what actually sits in it, WITHOUT the values ever going
    back to the model. Read deterministically here — code moves numbers, the LLM
    only ever named the column. Returns {} for fields whose header isn't found.
    """
    # Locate each header (by normalised label) to a (sheet, column index).
    index = {}
    for s in inspection.get("sheets", []):
        best = s.get("header_row_best_guess", 1)
        rows = s.get("first_rows", [])
        headers = s.get("headers", [])
        for col, label in enumerate(headers):
            key = _norm_header(label)
            if key and key not in index:
                index[key] = (best, rows, col)

    out = {}
    for field, header in _mapped_headers(mapping).items():
        loc = index.get(_norm_header(header))
        if not loc:
            out[field] = {"header": header, "samples": []}
            continue
        best, rows, col = loc
        samples = []
        for r in rows[best:]:  # rows AFTER the header row are data
            if col < len(r) and str(r[col]).strip():
                samples.append(str(r[col]).strip())
            if len(samples) >= max_samples:
                break
        out[field] = {"header": header, "samples": samples}
    return out


# --- /map : the single LLM call --------------------------------------------

def build_payload(inspection, sample_rows=3):
    """Minimal, non-sensitive payload sent to the model: per sheet, the guessed
    header row + up to `sample_rows` data rows. No full rate table ever leaves."""
    out = {"file": inspection["path"], "sheets": []}
    for s in inspection["sheets"]:
        best = s["header_row_best_guess"]
        data = s["first_rows"][best:best + sample_rows]
        out["sheets"].append({
            "name": s["name"],
            "header_row_candidates": s["header_row_candidates"][:3],
            "headers": s["headers"],
            "sample_data_rows": data,
        })
    return out


def _column_spec_schema():
    # A target field maps to: null | "Header" | {single:"Header"} | {split:"Header"}
    return {
        "oneOf": [
            {"type": "null"},
            {"type": "string"},
            {"type": "object", "properties": {"single": {"type": "string"}},
             "required": ["single"], "additionalProperties": False},
            {"type": "object", "properties": {"split": {"type": "string"}},
             "required": ["split"], "additionalProperties": False},
        ]
    }


def mapping_tool_schema():
    """input_schema for the forced tool call — the model returns a mapping.json."""
    return {
        "type": "object",
        "additionalProperties": False,
        "required": ["header_row", "output_prefix", "columns"],
        "properties": {
            "sheets": {"type": "array", "items": {"type": "string"},
                       "description": "Sheet names that each hold a term; omit for single-sheet/CSV."},
            "header_row": {"type": "integer", "minimum": 1,
                           "description": "1-based row holding the source headers."},
            "split_output_by_sheet": {"type": "boolean"},
            "output_prefix": {"type": "string"},
            "supplier": {"type": "string"},
            "term_labels": {"type": "object", "additionalProperties": {"type": "string"},
                            "description": "sheet name -> client-facing term label."},
            "category": {"type": "string"},
            "columns": {
                "type": "object",
                "additionalProperties": False,
                "properties": {f: _column_spec_schema() for f in TARGET_FIELDS},
            },
            "charge_basis": {"type": "object", "additionalProperties": {"type": "string"}},
            "db_lookup": {"type": "object"},
        },
    }


SYSTEM_PROMPT = (
    "You map a UK energy supplier quote's column headers onto RYE's fixed target "
    "schema. You NEVER transcribe or emit a rate, EAC or meter point — you only name "
    "which source column feeds each target field. Match by MEANING, not exact string.\n"
    "Target fields: " + ", ".join(TARGET_FIELDS) + ".\n"
    "Rules: use {\"single\": H} for a standard/anytime rate source and {\"split\": H} "
    "for day/night sources so single-vs-two-rate logic can tell them apart. If the "
    "supplier has no day/night columns, set dayRate/nightRate to {\"split\": \"__none__\"}. "
    "Band naming varies by supplier/SSC: map peak/on-peak to dayRate and off-peak to "
    "nightRate (they are the same two bands under different names). Use weekendRate "
    "(also a {\"split\": H}) ONLY when there is a genuinely distinct weekend column "
    "separate from day and night — not for peak/off-peak relabelling. "
    "capacityCharge vs kva are DIFFERENT and often confused: capacityCharge is a "
    "PRICE per unit of capacity (a column like 'KVA Charge (p/kVA/day)', 'Capacity "
    "Charge', 'Availability charge' — units of £/p per kVA per day), so map such a "
    "column to capacityCharge. kva is the agreed capacity QUANTITY itself (a plain "
    "kVA figure like 100 or 250, no price units); map a column to kva only when it "
    "holds that quantity. If a 'kVA charge' column is a price, it is capacityCharge, "
    "NOT kva — leave kva null (the agreed kVA usually comes from RYE's site reference). "
    "Set a field to null if the supplier does not provide it. Gas is single-rate: map its "
    "standard rate to unitRate. Pick the header_row that actually holds MPAN/EAC/rate "
    "labels, not a metadata/branding row. If the term is encoded in sheet names, list them "
    "in 'sheets' with split_output_by_sheet true and give client-facing term_labels."
)


def _client():
    try:
        import anthropic
    except ImportError:
        raise SystemExit(
            "anthropic SDK not installed. `pip install anthropic`. "
            "(Not needed for --dry-run or the deterministic path.)")
    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        raise SystemExit("ANTHROPIC_API_KEY not set — required for a live mapping call.")
    kwargs = {"api_key": key}
    base = os.environ.get("ANTHROPIC_BASE_URL")
    if base:
        kwargs["base_url"] = base  # e.g. Vercel AI Gateway endpoint
    return anthropic.Anthropic(**kwargs)


def propose_mapping(inspection, supplier=None, sample_rows=3, model=None):
    """The single LLM call. Returns a mapping dict (validated shape via tool use)."""
    payload = build_payload(inspection, sample_rows=sample_rows)
    model = model or os.environ.get("ANTHROPIC_MODEL", DEFAULT_MODEL)
    user = (
        (f"Supplier: {supplier}\n" if supplier else "")
        + "Inspect this quote and return the mapping via the emit_mapping tool.\n\n"
        + json.dumps(payload, ensure_ascii=False, indent=2)
    )
    client = _client()
    resp = client.messages.create(
        model=model,
        max_tokens=2000,
        system=SYSTEM_PROMPT,
        tools=[{
            "name": "emit_mapping",
            "description": "Return the mapping.json for this supplier quote.",
            "input_schema": mapping_tool_schema(),
        }],
        tool_choice={"type": "tool", "name": "emit_mapping"},
        messages=[{"role": "user", "content": user}],
    )
    for block in resp.content:
        if getattr(block, "type", None) == "tool_use" and block.name == "emit_mapping":
            mapping = dict(block.input)
            if supplier and "supplier" not in mapping:
                mapping["supplier"] = supplier
            mapping.setdefault("rate_logic", "auto")
            return mapping
    raise SystemExit("Model did not return an emit_mapping tool call.")


def autobuild(path, supplier=None, sample_rows=3, model=None):
    """inspect + propose -> ready-to-use mapping dict (called by process_quote --auto-map)."""
    inspection = inspect_file(path)
    return propose_mapping(inspection, supplier=supplier,
                           sample_rows=sample_rows, model=model)


def main(argv):
    p = argparse.ArgumentParser(description="Propose a mapping.json via Claude.")
    p.add_argument("source")
    p.add_argument("--supplier")
    p.add_argument("--sample-rows", type=int, default=3)
    p.add_argument("--dry-run", action="store_true",
                   help="print inspection + the exact request payload; make no API call")
    p.add_argument("--out", help="write the proposed mapping.json here")
    args = p.parse_args(argv)

    inspection = inspect_file(args.source)
    if args.dry_run:
        print("=== INSPECTION ===")
        print(json.dumps(inspection, ensure_ascii=False, indent=2))
        print("\n=== PAYLOAD THAT WOULD BE SENT TO CLAUDE (headers + samples only) ===")
        print(json.dumps(build_payload(inspection, args.sample_rows),
                         ensure_ascii=False, indent=2))
        print("\n(no API call made)")
        return
    mapping = propose_mapping(inspection, supplier=args.supplier,
                             sample_rows=args.sample_rows)
    out = json.dumps(mapping, ensure_ascii=False, indent=2)
    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(out)
        print(f"wrote {args.out}")
    else:
        print(out)


if __name__ == "__main__":
    main(sys.argv[1:])
