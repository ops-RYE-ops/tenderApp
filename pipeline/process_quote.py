#!/usr/bin/env python3
"""
RYE quote processor — deterministic extraction, no value invention.

Reads a supplier quote (xlsx/csv) plus a mapping.json that says which SOURCE
column feeds each TARGET field. All values are copied through verbatim (only
float->int tidy-up on identifiers and a decoration-stripping numeric parse on
rates). Nothing is computed or inferred here, so there is no room for
hallucination.

This is the Phase-0 extension of the original skill script. Two things are new,
both ADDED around the original logic (the value-moving core is untouched):

  1. Canonical JSON emission. Alongside the CSVs it always wrote, it now emits
     an `extractResult` document (schema/tender.schema.json #/$defs/extractResult):
     the sites[] discovered plus one quotes[] entry per term, with rates inlined
     as typed numbers. This is the shape /assemble stitches into a full tender.

  2. Optional Claude mapping call. With --auto-map it asks map_headers.py to
     inspect the file and propose the mapping (headers + a few sample rows only),
     instead of requiring a hand-written mapping.json. The LLM never sees full
     rate data; see map_headers.py.

Usage (backward compatible with the original positional form):
    python3 process_quote.py SOURCE mapping.json OUT_DIR [DB_CSV]

    # extras:
    #   --supplier "EDF"        supplier label for the canonical quotes[]
    #   --json-out PATH         where to write the extractResult JSON
    #                           (default: <OUT_DIR>/run-.../<prefix>.extract.json)
    #   --no-csv                skip the CSVs, emit canonical JSON only
    #   --auto-map [--map-out PATH]
    #                           build the mapping via Claude instead of reading
    #                           mapping.json (mapping arg then optional)
    #   --sample-rows N         sample rows sent to Claude for --auto-map (default 3)

mapping.json shape (produced by the skill's AI header-mapping step):
{
  "sheets": ["12m", "24m"],          # omit/empty for single-sheet or CSV input
  "header_row": 1,                    # 1-based row holding source headers
  "split_output_by_sheet": true,      # one CSV/quote per sheet (term)
  "output_prefix": "amorino",
  "supplier": "EDF",                  # optional; canonical supplier label
  "term_labels": { "12m": "12 months" },  # optional; sheet -> client-facing term
  "category": "fixed",                # optional; canonical offer category
  "charge_basis": { "meterCharge": "gbp/year" },  # optional; per-quote override
  "columns": {                        # TARGET field -> source header (or null)
      "siteName":        "Address",
      "mpxn":            "Meter Point",
      "updatedEac":      "Industry Consumption (EAC/AQ)",
      "supplyStartDate": "Start Date",
      "unitRate":        {"single": "Standard"},
      "dayRate":         {"split": "Day"},
      "nightRate":       {"split": "Night"},
      "standingCharge":  "Standing Charge",
      "capacityCharge":  null,
      "networkCharge":   "DUoS",
      "meterCharge":     "Meter Charge",
      "kva":             "kVa Capacity"
  },
  "rate_logic": "auto",
  "db_lookup": { "mpxn_col": "mpxn", "name_col": "siteName" }
}

The TARGET header order is fixed by TARGET_HEADERS below to match RYE's internal CSV.
"""
import argparse
import csv
import datetime
import json
import os
import re
import sys

from rye_quote_core import TARGET_FIELDS as TARGET_HEADERS, parse_num

# Fields that live on a canonical quote LINE (rates/charges), in schema order.
LINE_RATE_FIELDS = [
    "unitRate", "dayRate", "nightRate", "standingCharge",
    "capacityCharge", "networkCharge", "meterCharge",
]


def slugify(s):
    """Filesystem-safe sheet label: 'Fixed 12 Months' -> 'fixed-12-months'."""
    s = str(s).strip().lower()
    s = re.sub(r"[%/\\]+", " ", s)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "term"


def clean_id(v):
    """Meter points / numeric IDs often arrive as floats (e.g. 1200035438587.0)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def clean_val(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def build_name_lookup(db_csv, mpxn_col, name_col):
    """mpxn -> our site name, from RYE's DB export. Keys are normalised mpxns."""
    lookup = {}
    with open(db_csv, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if mpxn_col not in reader.fieldnames or name_col not in reader.fieldnames:
            raise SystemExit(
                f"DB CSV must contain columns '{mpxn_col}' and '{name_col}'. "
                f"Found: {reader.fieldnames}"
            )
        for r in reader:
            key = clean_id(r.get(mpxn_col))
            name = (r.get(name_col) or "").strip()
            if key and name:
                lookup[key] = name
    return lookup


def resolve_constants(mapping, src_path):
    """Build {target: value} from mapping['constants'].

    Each value is either a literal (used as-is) or {"cell": "I13"} which reads
    that cell from the source sheet (xlsx only). Used for fields the per-row
    table doesn't carry, e.g. a single contract start date in a summary block.
    """
    spec = mapping.get("constants")
    if not spec:
        return {}
    out = {}
    ws = None
    for target, v in spec.items():
        if isinstance(v, dict) and "cell" in v:
            if ws is None:
                import openpyxl
                wb = openpyxl.load_workbook(src_path, data_only=True)
                sheets = mapping.get("sheets") or [None]
                ws = wb[sheets[0]] if sheets[0] else wb[wb.sheetnames[0]]
            out[target] = ws[v["cell"]].value
        else:
            out[target] = v
    return out


def load_rows(path, sheet, header_row):
    """Return list of dicts keyed by source header. Works for xlsx and csv."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            rows = [tuple(r) for r in csv.reader(f)]
    if not rows:
        return []
    hdr = [clean_val(c) for c in rows[header_row - 1]]
    out = []
    for raw in rows[header_row:]:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        rec = {}
        for i, h in enumerate(hdr):
            rec[h] = raw[i] if i < len(raw) else None
        out.append(rec)
    return out


def resolve(spec, rec):
    """Resolve one target column spec against a source record.
    Returns (value, kind) where kind is 'single', 'split', or 'plain'.
    """
    if spec is None:
        return None, "plain"
    if isinstance(spec, str):
        return rec.get(spec), "plain"
    if isinstance(spec, dict):
        if "single" in spec:
            return rec.get(spec["single"]), "single"
        if "split" in spec:
            return rec.get(spec["split"]), "split"
    return None, "plain"


def process_rows(records, mapping, name_lookup=None, constants=None):
    cols = mapping["columns"]
    constants = constants or {}
    out_rows = []
    unmatched = []
    for rec in records:
        # Determine rate mode for this row: split (day/night) vs single.
        day_spec = cols.get("dayRate")
        day_val, _ = resolve(day_spec, rec)
        night_spec = cols.get("nightRate")
        night_val, _ = resolve(night_spec, rec)
        is_split = (day_val not in (None, "")) or (night_val not in (None, ""))

        row = {}
        for target in TARGET_HEADERS:
            spec = cols.get(target)
            val, kind = resolve(spec, rec)
            # Apply single/split exclusivity per the agreed rate mapping.
            if kind == "single" and is_split:
                val = None            # split row -> leave unitRate blank
            if kind == "split" and not is_split:
                val = None            # single row -> leave day/night blank
            row[target] = clean_id(val) if target == "mpxn" else clean_val(val)
        # Apply fixed constants (e.g. a global start date from the summary block)
        # to any target the per-row table didn't fill.
        for target, cval in constants.items():
            if target in TARGET_HEADERS and not row.get(target):
                row[target] = clean_val(cval)
        # Override siteName with OUR DB name, matched on mpxn (the join key).
        if name_lookup:
            our_name = name_lookup.get(row.get("mpxn", ""))
            if our_name:
                row["siteName"] = our_name
            else:
                unmatched.append(row.get("mpxn", ""))
        out_rows.append(row)
    return out_rows, unmatched


def unique_path(path):
    """Return a path that does not exist yet, suffixing -2, -3, ... if needed.
    Prevents a new run from overwriting an earlier quote's output."""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 2
    while os.path.exists(f"{base}-{n}{ext}"):
        n += 1
    return f"{base}-{n}{ext}"


def write_csv(rows, path, overwrite=False):
    if not overwrite:
        path = unique_path(path)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=TARGET_HEADERS)
        w.writeheader()
        w.writerows(rows)
    return path


# --- Canonical JSON emission (Phase-0 addition) ---------------------------
# Everything below turns the already-extracted TARGET-schema rows into the
# canonical `extractResult` shape. It reads only the rows produced by
# process_rows above — it never re-reads the source or re-decides a value — so
# the CSV and the JSON are guaranteed to agree.

def row_to_line(row):
    """One TARGET-schema row -> a canonical quote line (rates as numbers|null).

    mpxn stays a string. supplyStartDate passes through as a string (already
    normalised to YYYY-MM-DD by clean_val). Every rate/charge is run through the
    shared parse_num so the stored number is exactly what the cost engine sees.
    Fields left blank by the mapping become null (not quoted).
    """
    line = {"mpxn": row.get("mpxn", "")}
    sd = (row.get("supplyStartDate") or "").strip()
    line["supplyStartDate"] = sd or None
    for f in LINE_RATE_FIELDS:
        line[f] = parse_num(row.get(f))
    return line


def row_to_site(row, eac_source):
    """One TARGET-schema row -> a canonical site entry (the meter facts)."""
    return {
        "mpxn": row.get("mpxn", ""),
        "site_name": (row.get("siteName") or "").strip() or row.get("mpxn", ""),
        "eac": parse_num(row.get("updatedEac")),
        "kva": parse_num(row.get("kva")),
        "eac_source": eac_source,
    }


def rows_to_quote(rows, supplier, term, category=None, charge_basis=None):
    """Build one canonical quotes[] entry from a term's TARGET-schema rows."""
    quote = {"supplier": supplier}
    if term:
        quote["term"] = term
    if category:
        quote["category"] = category
    if charge_basis:
        quote["charge_basis"] = charge_basis
    quote["lines"] = [row_to_line(r) for r in rows if r.get("mpxn")]
    return quote


def term_for_sheet(mapping, sheet, split_out):
    """Client-facing term label for a sheet/term."""
    if not (split_out and sheet):
        return mapping.get("term", "")
    labels = mapping.get("term_labels") or {}
    return labels.get(sheet, str(sheet))


# ---------------------------------------------------------------------------


def run(src, mapping, out_dir, db_csv=None, supplier=None, emit_csv=True,
        json_out=None):
    """Core pipeline. Returns (written_csvs, extract_result, all_unmatched).

    extract_result is the canonical `extractResult` dict. Kept as a function so
    the Vercel /extract endpoint can import and call it directly rather than
    shelling out — matching the spec's 'functions import the scripts' principle.
    """
    if mapping.get("run_subdir", True) and os.environ.get("QP_NO_RUN_SUBDIR") != "1":
        stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        out_dir = os.path.join(out_dir, f"run-{stamp}")
    os.makedirs(out_dir, exist_ok=True)

    name_lookup = None
    if db_csv:
        dbl = mapping.get("db_lookup", {})
        name_lookup = build_name_lookup(
            db_csv, dbl.get("mpxn_col", "mpxn"), dbl.get("name_col", "siteName")
        )
        print(f"loaded {len(name_lookup)} site names from {db_csv}")

    constants = resolve_constants(mapping, src)
    if constants:
        print(f"applying constants: {constants}")

    sheets = mapping.get("sheets") or [None]
    prefix = mapping.get("output_prefix", "processed")
    split_out = mapping.get("split_output_by_sheet", True)
    header_row = mapping.get("header_row", 1)
    supplier = supplier or mapping.get("supplier") or prefix
    category = mapping.get("category")
    charge_basis = mapping.get("charge_basis") or None
    # EAC/kVA here come off the supplier quote, so record that provenance. The
    # backend can promote these to 'db' once site reference data is joined in.
    eac_source = "quote"

    written = []
    combined = []
    all_unmatched = set()
    canonical_quotes = []
    sites_by_mpxn = {}

    def collect_sites(rows):
        for r in rows:
            m = r.get("mpxn")
            if m and m not in sites_by_mpxn:
                sites_by_mpxn[m] = row_to_site(r, eac_source)

    for sheet in sheets:
        records = load_rows(src, sheet, header_row)
        rows, unmatched = process_rows(records, mapping, name_lookup, constants)
        all_unmatched.update(u for u in unmatched if u)
        collect_sites(rows)
        term = term_for_sheet(mapping, sheet, split_out)
        if split_out and sheet:
            canonical_quotes.append(
                rows_to_quote(rows, supplier, term, category, charge_basis))
            if emit_csv:
                p = os.path.join(out_dir, f"{prefix}-{slugify(sheet)}.csv")
                p = write_csv(rows, p)
                written.append((p, len(rows)))
        else:
            combined.extend(rows)

    if combined or not canonical_quotes:
        term = mapping.get("term", "")
        canonical_quotes.append(
            rows_to_quote(combined, supplier, term, category, charge_basis))
        if emit_csv and (combined or not written):
            p = os.path.join(out_dir, f"{prefix}.csv")
            p = write_csv(combined, p)
            written.append((p, len(combined)))

    extract_result = {
        "sites": list(sites_by_mpxn.values()),
        "quotes": canonical_quotes,
    }

    # Write the canonical JSON (non-overwriting, same discipline as the CSVs).
    if json_out is None:
        json_out = os.path.join(out_dir, f"{prefix}.extract.json")
    json_out = unique_path(json_out)
    with open(json_out, "w", encoding="utf-8") as f:
        json.dump(extract_result, f, indent=2, ensure_ascii=False)
    extract_result["_json_path"] = json_out

    return written, extract_result, all_unmatched


def parse_args(argv):
    p = argparse.ArgumentParser(add_help=True, description="RYE deterministic quote extractor")
    p.add_argument("source", help="supplier quote (xlsx/csv)")
    p.add_argument("mapping", nargs="?", help="mapping.json (optional with --auto-map)")
    p.add_argument("out_dir", help="output directory")
    p.add_argument("db_csv", nargs="?", help="optional RYE DB export for site-name lookup")
    p.add_argument("--supplier", help="supplier label for the canonical quotes[]")
    p.add_argument("--json-out", dest="json_out", help="path for the extractResult JSON")
    p.add_argument("--no-csv", action="store_true", help="emit canonical JSON only")
    p.add_argument("--auto-map", action="store_true",
                   help="build the mapping via Claude (map_headers.py) instead of reading mapping.json")
    p.add_argument("--map-out", dest="map_out", help="where to save the auto-built mapping.json")
    p.add_argument("--sample-rows", type=int, default=3,
                   help="sample rows sent to Claude for --auto-map (default 3)")
    return p.parse_args(argv)


def main(argv):
    args = parse_args(argv)

    if args.auto_map:
        # Lazy import so the deterministic path has zero LLM dependencies.
        import map_headers
        mapping = map_headers.autobuild(
            args.source,
            supplier=args.supplier,
            sample_rows=args.sample_rows,
        )
        if args.map_out:
            with open(args.map_out, "w", encoding="utf-8") as f:
                json.dump(mapping, f, indent=2, ensure_ascii=False)
            print(f"auto-built mapping -> {args.map_out}")
    else:
        if not args.mapping:
            raise SystemExit("mapping.json is required unless --auto-map is given")
        with open(args.mapping) as f:
            mapping = json.load(f)

    written, extract_result, all_unmatched = run(
        args.source, mapping, args.out_dir, db_csv=args.db_csv,
        supplier=args.supplier, emit_csv=not args.no_csv, json_out=args.json_out,
    )

    for p, n in written:
        print(f"wrote {n} rows -> {p}")
    print(f"wrote canonical JSON -> {extract_result['_json_path']} "
          f"({len(extract_result['sites'])} sites, {len(extract_result['quotes'])} quote-term(s))")
    if all_unmatched:
        print(f"WARNING: {len(all_unmatched)} mpxn(s) had no DB site-name match "
              f"(kept supplier address): {', '.join(sorted(all_unmatched))}")


if __name__ == "__main__":
    main(sys.argv[1:])
